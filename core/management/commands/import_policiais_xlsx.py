from __future__ import annotations

import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from core.models import Policial


def norm_text(value: str) -> str:
    text = (value or "").strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_matricula(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    text = text.replace(" ", "")
    if text.endswith(".0"):
        inteiro = text[:-2]
        if inteiro.isdigit():
            text = inteiro

    return text


class Command(BaseCommand):
    help = "Importa policiais (servidores) de planilha XLSX com upsert por matricula."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="servidores.xlsx",
            help="Caminho da planilha XLSX (padrao: servidores.xlsx na raiz do workspace).",
        )
        parser.add_argument(
            "--sheet",
            default="",
            help="Nome da aba a importar. Se vazio, usa a primeira aba.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Aplica alteracoes. Sem essa flag, roda em dry-run.",
        )

    def _resolve_file(self, file_arg: str) -> Path:
        candidate = Path(file_arg)
        if candidate.is_absolute() and candidate.exists():
            return candidate

        workspace_root = Path(__file__).resolve().parents[4]
        candidate = workspace_root / file_arg
        if candidate.exists():
            return candidate

        raise FileNotFoundError(f"Arquivo nao encontrado: {file_arg}")

    @staticmethod
    def _header_map(header_row):
        mapping = {}
        for idx, raw in enumerate(header_row):
            key = norm_text(cell_value(raw))
            if key:
                mapping[key] = idx
        return mapping

    @staticmethod
    def _pick(row, header_map, *aliases):
        for alias in aliases:
            idx = header_map.get(norm_text(alias))
            if idx is None:
                continue
            if idx < len(row):
                return cell_value(row[idx])
        return ""

    def handle(self, *args, **options):
        dry_run = not options["apply"]
        file_path = self._resolve_file(options["file"])

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet_name = options["sheet"] or wb.sheetnames[0]
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Aba '{sheet_name}' nao existe. Abas disponiveis: {', '.join(wb.sheetnames)}")

            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                raise ValueError("Planilha vazia.")

            header_map = self._header_map(header_row)

            required = ["matricula", "nome"]
            missing = [col for col in required if col not in header_map]
            if missing:
                raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(missing)}")

            counters = {
                "linhas_lidas": 0,
                "ignoradas_sem_chave": 0,
                "criadas": 0,
                "atualizadas": 0,
                "inalteradas": 0,
            }

            existing = {norm_text(normalize_matricula(pol.matricula)): pol for pol in Policial.objects.all()}
            seen_sheet_keys = set()

            with transaction.atomic():
                for row in rows:
                    counters["linhas_lidas"] += 1

                    matricula_raw = self._pick(row, header_map, "matricula", "matricula funcional", "matricula_funcional")
                    nome = self._pick(row, header_map, "nome", "nome completo", "servidor", "policial")
                    cpf = self._pick(row, header_map, "cpf")
                    cargo = self._pick(
                        row,
                        header_map,
                        "cargo",
                        "cargo atual",
                        "funcao",
                        "funcao/cargo",
                    )
                    depto = self._pick(row, header_map, "depto", "departamento")
                    if not depto:
                        depto = self._pick(
                            row,
                            header_map,
                            "departamento (atual)",
                            "departamento(atual)",
                            "departamento atual",
                            "und. departamento (atual)",
                            "und departamento (atual)",
                            "und. departamento atual",
                            "und departamento atual",
                        )
                    lotacao = self._pick(
                        row,
                        header_map,
                        "lotacao",
                        "lotacao/unidade",
                        "und. lotacao (atual)",
                        "und lotacao (atual)",
                        "und. lotacao atual",
                        "und lotacao atual",
                        "delegacia",
                        "unidade",
                    )
                    tel = self._pick(row, header_map, "telefone", "tel", "celular")
                    email = self._pick(row, header_map, "email", "e-mail")
                    obs = self._pick(row, header_map, "observacoes", "observacao", "obs")

                    matricula = normalize_matricula(matricula_raw)

                    if not matricula or not nome:
                        counters["ignoradas_sem_chave"] += 1
                        continue

                    key = norm_text(matricula)
                    if key in seen_sheet_keys:
                        counters["inalteradas"] += 1
                        continue
                    seen_sheet_keys.add(key)

                    pol = existing.get(key)

                    if pol is None:
                        Policial.objects.create(
                            matricula=matricula,
                            nome=nome,
                            cpf=cpf,
                            cargo=cargo,
                            depto=depto,
                            lotacao=lotacao,
                            tel=tel,
                            email=email,
                            obs=obs,
                        )
                        existing[key] = Policial(
                            matricula=matricula,
                            nome=nome,
                            cpf=cpf,
                            cargo=cargo,
                            depto=depto,
                            lotacao=lotacao,
                            tel=tel,
                            email=email,
                            obs=obs,
                        )
                        counters["criadas"] += 1
                        continue

                    changed = False
                    if pol.nome != nome:
                        pol.nome = nome
                        changed = True
                    if pol.cpf != cpf:
                        pol.cpf = cpf
                        changed = True
                    if pol.cargo != cargo:
                        pol.cargo = cargo
                        changed = True
                    if pol.depto != depto:
                        pol.depto = depto
                        changed = True
                    if pol.lotacao != lotacao:
                        pol.lotacao = lotacao
                        changed = True
                    if pol.tel != tel:
                        pol.tel = tel
                        changed = True
                    if pol.email != email:
                        pol.email = email
                        changed = True
                    if pol.obs != obs:
                        pol.obs = obs
                        changed = True

                    if changed:
                        pol.save(
                            update_fields=[
                                "nome",
                                "cpf",
                                "cargo",
                                "depto",
                                "lotacao",
                                "tel",
                                "email",
                                "obs",
                                "atualizado_em",
                            ]
                        )
                        counters["atualizadas"] += 1
                    else:
                        counters["inalteradas"] += 1

                if dry_run:
                    transaction.set_rollback(True)

            mode_msg = "DRY-RUN (sem persistir)" if dry_run else "APLICADO"
            self.stdout.write(self.style.WARNING(f"Importacao de policiais: {mode_msg}"))
            for key, value in counters.items():
                self.stdout.write(f"- {key}: {value}")
        finally:
            wb.close()
