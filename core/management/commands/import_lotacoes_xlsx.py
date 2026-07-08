from __future__ import annotations

import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from core.models import Lotacao


def norm_text(value: str) -> str:
    text = (value or "").strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Importa lotacoes a partir de planilha XLSX com colunas DELEGACIA/DEPARTAMENTO/SECCIONAL/AREA DE ATUACAO/AIS."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="circuscricao.xlsx",
            help="Caminho da planilha XLSX (padrao: circuscricao.xlsx na raiz do workspace).",
        )
        parser.add_argument(
            "--sheet",
            default="",
            help="Nome da aba a importar. Se vazio, usa a primeira aba.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Aplica alteracoes. Sem essa flag, roda em dry-run.",
        )

    def _resolve_file(self, file_arg: str) -> Path:
        candidate = Path(file_arg)
        if candidate.is_absolute() and candidate.exists():
            return candidate

        # backend/ -> workspace root
        workspace_root = Path(__file__).resolve().parents[4]
        candidate = workspace_root / file_arg
        if candidate.exists():
            return candidate

        raise FileNotFoundError(f"Arquivo nao encontrado: {file_arg}")

    @staticmethod
    def _header_map(header_row):
        mapping = {}
        for idx, raw in enumerate(header_row):
            key = norm_text(cell_value(raw))
            if key:
                mapping[key] = idx
        return mapping

    @staticmethod
    def _pick(row, header_map, *aliases):
        for alias in aliases:
            idx = header_map.get(alias)
            if idx is None:
                continue
            if idx < len(row):
                return cell_value(row[idx])
        return ""

    def handle(self, *args, **options):
        dry_run = not options["apply"]
        file_path = self._resolve_file(options["file"])

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_name = options["sheet"] or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' nao existe. Abas disponiveis: {', '.join(wb.sheetnames)}")

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise ValueError("Planilha vazia.")

        header_map = self._header_map(header_row)

        required = ["delegacia", "departamento"]
        missing = [col for col in required if col not in header_map]
        if missing:
            raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(missing)}")

        counters = {
            "linhas_lidas": 0,
            "ignoradas_sem_chave": 0,
            "criadas": 0,
            "atualizadas": 0,
            "inalteradas": 0,
        }

        existing = {
            (norm_text(lot.depto), norm_text(lot.nome)): lot
            for lot in Lotacao.objects.all()
        }

        with transaction.atomic():
            for row in rows:
                counters["linhas_lidas"] += 1

                delegacia = self._pick(row, header_map, "delegacia")
                departamento = self._pick(row, header_map, "departamento")
                seccional = self._pick(row, header_map, "seccional")
                area = self._pick(row, header_map, "area de atuacao", "area de atuação")
                ais = self._pick(row, header_map, "ais")

                if not delegacia or not departamento:
                    counters["ignoradas_sem_chave"] += 1
                    continue

                cidade = seccional or area or ais or "N/D"
                resp = area

                endereco_parts = []
                if seccional:
                    endereco_parts.append(f"Seccional: {seccional}")
                if area:
                    endereco_parts.append(f"Area: {area}")
                if ais:
                    endereco_parts.append(f"AIS: {ais}")
                endereco = " | ".join(endereco_parts)

                key = (norm_text(departamento), norm_text(delegacia))
                lot = existing.get(key)

                if lot is None:
                    Lotacao.objects.create(
                        depto=departamento,
                        nome=delegacia,
                        cidade=cidade,
                        resp=resp,
                        end=endereco,
                    )
                    counters["criadas"] += 1
                    continue

                changed = False
                if lot.cidade != cidade:
                    lot.cidade = cidade
                    changed = True
                if resp and lot.resp != resp:
                    lot.resp = resp
                    changed = True
                if endereco and lot.end != endereco:
                    lot.end = endereco
                    changed = True

                if changed:
                    lot.save(update_fields=["cidade", "resp", "end", "atualizado_em"])
                    counters["atualizadas"] += 1
                else:
                    counters["inalteradas"] += 1

            if dry_run:
                transaction.set_rollback(True)

        mode_msg = "DRY-RUN (sem persistir)" if dry_run else "APLICADO"
        self.stdout.write(self.style.WARNING(f"Importacao de lotacoes: {mode_msg}"))
        for k, v in counters.items():
            self.stdout.write(f"- {k}: {v}")
